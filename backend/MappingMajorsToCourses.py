from typing import Any

import sys
sys.path.append("../")

import backend.ClassAssignment as ClassAssignment
import backend.Student as Student
import backend.Course as Course
import openpyxl


# class Major:
#     def __init__(self, major_list):
#         self.major_list = major_list
#
#     def get_major_list(self):
#         return self.major_list


def get_major_list(students, courses):
    """
    get students' major interest from the main tab of the spreadsheet, and a list of the names of all courses.
    @param students: a list containing all students as objects of Student.py
    @param courses: a list containing all courses as objects of Course.py
    returns a list of all major that students are interested in.
            a list of names of all classes.
    """
    major_list = []
    classes_list = []

    for s in students:
        mj1 = s.get_student_detail("Maj_Int_1")
        mj2 = s.get_student_detail("Maj_Int_2")
        mj3 = s.get_student_detail("Maj_Int_3")
        mj4 = s.get_student_detail("Maj_Int_4")
        if mj1 not in major_list:
            if mj1 is not None:
                major_list.append(mj1)
        if mj2 not in major_list:
            if mj2 is not None:
                major_list.append(mj2)
        if mj3 not in major_list:
            if mj3 is not None:
                major_list.append(mj3)
        if mj4 not in major_list:
            if mj4 is not None:
                major_list.append(mj4)

    for c in courses:
        classes_list.append(c.getName())

    # Major called 'Interdepartmental (combining the areas of study listed for the next two questions)' needs to be
    # accounted for algorithm wise.
    major_list.remove('Interdepartmental (combining the areas of study listed for the next two questions)')

    return major_list, classes_list


def add_class_to_major(major_class_dict, a_class, a_major):
    """
    add a course to a major
    @param major_class_dict: the mapped dictionary.
    @param a_class: a string of a class's name.
    @param a_major: a string of a major.
    """
    major_class_dict[a_major].append(a_class)


def map_courses_to_major(students, courses):
    """
    map available classes to the major that students are interested in (from maintab).
    @param students: a list containing all students as objects of Student.py
    @param courses: a list containing all courses as objects of Course.py
    returns a dictionary containing majors as keys,with values that are classes that fit those majors.
    """

    majors, classes = get_major_list(students, courses)
    majors.sort()
    courses_to_majors_dict = {}
    for mj in majors:
        courses_to_majors_dict[mj] = []

    for a_class in classes:
        if a_class[0] == "*":
            dept = a_class[1:4]
        else:
            dept = a_class[0:3]
            
        if dept == 'AAH':
            courses_to_majors_dict['Visual Arts (Art History Concentration)'].append(a_class)
            courses_to_majors_dict['Visual Arts (Art History/Studio Arts Dual Concentration)'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'ADA':
            courses_to_majors_dict['American Studies (I)'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'AMU':
            courses_to_majors_dict['Music'].append(a_class)
        if a_class == 'ANT-227-01 Policing the Americas':
            courses_to_majors_dict['American Studies (I)'].append(a_class)
        if dept == 'ANT':
            courses_to_majors_dict['Anthropology'].append(a_class)
        if dept == 'ARB':
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'AST':
            courses_to_majors_dict['Astronomy'].append(a_class)
            courses_to_majors_dict['Astronomy (Lab Science)'].append(a_class)
        if dept == 'ATH':
            courses_to_majors_dict['Theater'].append(a_class)
        if dept == 'AVA':
            courses_to_majors_dict['Visual Arts (Art History Concentration)'].append(a_class)
            courses_to_majors_dict['Visual Arts (Studio Fine Arts Concentration)'].append(a_class)
            courses_to_majors_dict['Visual Arts (Art History/Studio Arts Dual Concentration)'].append(a_class)
        if dept == 'BIO':
            courses_to_majors_dict['Biochemistry'].append(a_class)
            courses_to_majors_dict['Biochemistry (Lab Science)'].append(a_class)
            courses_to_majors_dict['Bioengineering'].append(a_class)
            courses_to_majors_dict['Biology'].append(a_class)
            courses_to_majors_dict['Biology (Lab Science)'].append(a_class)
            courses_to_majors_dict['Biomedical Engineering'].append(a_class)
            courses_to_majors_dict['Neuroscience (I)'].append(a_class)
        if dept == 'CHM':
            courses_to_majors_dict['Chemistry'].append(a_class)
            courses_to_majors_dict['Chemistry (Lab Science)'].append(a_class)
            courses_to_majors_dict['Biochemistry'].append(a_class)
            courses_to_majors_dict['Biochemistry (Lab Science)'].append(a_class)
        if dept == 'CHN':
            courses_to_majors_dict['Modern Languages: Chinese'].append(a_class)
        if dept == 'CLS':
            courses_to_majors_dict['Classics'].append(a_class)
        if dept == 'CSC':
            courses_to_majors_dict['Computer Science'].append(a_class)
            courses_to_majors_dict['Electrical Engineering'].append(a_class)
            courses_to_majors_dict['Computer Engineering'].append(a_class)
            courses_to_majors_dict['Mechanical Engineering'].append(a_class)
            courses_to_majors_dict['Undecided Engineering'].append(a_class)
        if dept == 'ECO':
            courses_to_majors_dict['Economics'].append(a_class)
            courses_to_majors_dict['Managerial Economics'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
            courses_to_majors_dict['*Exploring Social Science majors'].append(a_class)
        if dept == 'EGL':
            courses_to_majors_dict['English'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'ENS':
            courses_to_majors_dict['Environmental Science (I)'].append(a_class)
            courses_to_majors_dict['Environmental Policy (I)'].append(a_class)
        if dept == "ESC":
            courses_to_majors_dict['Bioengineering'].append(a_class)
            courses_to_majors_dict['Biomedical Engineering'].append(a_class)
            courses_to_majors_dict['Electrical Engineering'].append(a_class)
            courses_to_majors_dict['Computer Engineering'].append(a_class)
            courses_to_majors_dict['Mechanical Engineering'].append(a_class)
            courses_to_majors_dict['Undecided Engineering'].append(a_class)
        if dept == 'FRN':
            courses_to_majors_dict['Modern Languages: French and Francophone Studies'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'GEO':
            courses_to_majors_dict['Geology'].append(a_class)
            courses_to_majors_dict['Geology (Lab Science)'].append(a_class)
        if dept == 'GER':
            courses_to_majors_dict['Modern Languages: German Studies'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'GRK':
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
            courses_to_majors_dict['Classics'].append(a_class)
        if dept == 'GSW':
            courses_to_majors_dict['Gender, Sexuality, and Women_s Studies (I)'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'HEB':
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'HST':
            courses_to_majors_dict['History'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'JPN':
            courses_to_majors_dict['Asian Studies (I)'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'LAT':
            courses_to_majors_dict['Classics'].append(a_class)
        if dept == 'MIN':
            courses_to_majors_dict['Environmental Science (I)'].append(a_class)
            courses_to_majors_dict['Environmental Policy (I)'].append(a_class)
            courses_to_majors_dict['*Exploring Social Science majors'].append(a_class)
            courses_to_majors_dict['Sociology'].append(a_class)
        if a_class == 'MLT-200-01 Modern Chinese Literature':
            courses_to_majors_dict['Modern Languages: Chinese'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if a_class == 'MLT-260-01 Vampire As Other E Eur/Am Cult':
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'MTH':
            courses_to_majors_dict['Mathematics'].append(a_class)
        if dept == 'PHL':
            courses_to_majors_dict['Philosophy'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if "PHY-100" in a_class:
            courses_to_majors_dict['Physics'].append(a_class)
            courses_to_majors_dict['Physics (Lab Science)'].append(a_class)
            courses_to_majors_dict['Astronomy'].append(a_class)
            courses_to_majors_dict['Astronomy (Lab Science)'].append(a_class)
        if dept == 'PHY':
            courses_to_majors_dict['Physics'].append(a_class)
            courses_to_majors_dict['Physics (Lab Science)'].append(a_class)
            courses_to_majors_dict['Undecided Engineering'].append(a_class)
        if dept == 'PSC':
            courses_to_majors_dict['*Exploring Social Science majors'].append(a_class)
            courses_to_majors_dict['Political Science'].append(a_class)
        if dept == 'PSY':
            courses_to_majors_dict['Psychology'].append(a_class)
            courses_to_majors_dict['Neuroscience (I)'].append(a_class)
        if dept == 'REL':
            courses_to_majors_dict['Religious Studies'].append(a_class)
            courses_to_majors_dict['*Exploring Arts and Humanities majors'].append(a_class)
        if dept == 'RUS':
            courses_to_majors_dict['Modern Languages: Russian (ID)'].append(a_class)
            courses_to_majors_dict['Russian and East European Studies (I)'].append(a_class)
        if (dept == "SMT") or ("HST-138" in a_class or "HST-293" in a_class
                               or "PHL-232" in a_class or "SOC-228" in a_class):
            courses_to_majors_dict['Science, Medicine, and Technology in Culture (I)'].append(a_class)
        if dept == 'SOC':
            courses_to_majors_dict['Sociology'].append(a_class)
            courses_to_majors_dict['Political Science'].append(a_class)
            courses_to_majors_dict['*Exploring Social Science majors'].append(a_class)
        if dept == 'SPN':
            courses_to_majors_dict['Modern Languages: Spanish and Hispanic Studies'].append(a_class)
            courses_to_majors_dict['Latin American and Caribbean Studies (I)'].append(a_class)
        if dept == 'STA':
            courses_to_majors_dict['Mathematics'].append(a_class)
            courses_to_majors_dict['**Exploring many potential majors'].append(a_class)
        if dept == 'HCX':
            pass

    print('_________________________________________')
    print('CAUTION: THESE MAJORS ARE HAVING ONLY ONE CLASS')
    for a_major in courses_to_majors_dict:
        if len(courses_to_majors_dict[a_major]) <= 1:
            print(a_major)
    print('_________________________________________')
    return courses_to_majors_dict


def _write_to_workbook(major_course_dict):
    """
    export a a xlsx file with mapping of courses to majors.
    @param major_course_dict: the mapped dictionary.
    """
    workbook = openpyxl.load_workbook(
        filename='Copy of Class of 2026 (Fall 2022) Advisor and Registration Assignments - Anonymized.xlsx')
    worksheet = workbook['Majors-Courses']
    data = []
    for a_major in major_course_dict:
        to_append = [a_major, major_course_dict[a_major]]
        data.append(to_append)

    # data = [['a',[1,1]], ['b',[2,2,2]]]

    for x in range(len(data)):
        worksheet.cell(row=x + 1, column=1, value=data[x][0])
        for y in range(len(data[x][1])):
            j = y + 2
            worksheet.cell(row=x + 1, column=j, value=data[x][1][y])
    workbook.save('Copy of Class of 2026 (Fall 2022) Advisor and Registration Assignments - Anonymized.xlsx')


if __name__ == '__main__':
    manager = ClassAssignment.create_manager()
    print("Data import complete")

    students = manager.get_students()
    courses = manager.get_courses()

    courses_to_majors_dict = map_courses_to_major(students, courses)
    # add_class_to_major(courses_to_majors_dict,'ANT-227-01 Policing the Americas','American Studies (I)')
    _write_to_workbook(courses_to_majors_dict)

    # String representation of the dictionary
    for a_major in courses_to_majors_dict:
        print(a_major + ": ", end='')
        for a_class in courses_to_majors_dict[a_major]:
            print(a_class + ', ', end='')
        print()
        print('__________________________________________________')
