#TODO: write methods appear to only save the first one run. Check if openpyxl needs to save/close a workbook once done with it?

import sys
import openpyxl
import pandas as pd
import numpy as np
from pprint import pprint
import datetime
from datetime import time

FILE_NAME = '../input_files/Copy of Class of 2026 (Fall 2022) Advisor and Registration Assignments - Anonymized.xlsx'

open_dfs = {}

err_log = open("data_errors.log", "w")


def open_if_not(filename=FILE_NAME):
    """Opens Excel file and extracts all sheets, caching the result for performance."""
    
    if filename not in open_dfs.keys():
        xlsx = pd.ExcelFile(filename)
        dfs = {}
        for sheet in xlsx.sheet_names:
            dfs[sheet] = pd.read_excel(xlsx, sheet)
            dfs[sheet] = dfs[sheet].replace(np.nan, None)
        open_dfs[filename] = dfs
    return open_dfs[filename]


def read_workbook(worksheet_name, filename=FILE_NAME):
    """
    reads a worksheet from an Excel file, and returns it as a list of rows
    [[contents of row 1], [contents of row 2], ..., [contents of last row with data]]
    """

    worksheet = open_if_not(filename)[worksheet_name]
    (max_row, max_col) = worksheet.shape
    
    dataset = [list(worksheet.columns)]
    for r in range(max_row):
        row = []
        for c in range(max_col):
            row.append(worksheet.iat[r, c])
        dataset.append(row)

    dataset_cleaned_transposed = []
    transposed_dataset = [list(i) for i in zip(*dataset)]
    c = 0
    for col in transposed_dataset:
        dataset_cleaned_transposed.append(clean_types(col, worksheet.columns[c]))
        c += 1

    dataset_cleaned = [list(i) for i in zip(*dataset_cleaned_transposed)]

    return dataset_cleaned


def read_workbook_row(worksheet_name, row_num, filename=FILE_NAME):
    """
    reads a single row of a worksheet from an Excel file and returns it as a list, with each list item
    representing one cell from the row
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param row_num: the number of the row to read (with the first row being row 1).
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """

    worksheet = open_if_not(filename)[worksheet_name]
    (max_row, max_col) = worksheet.shape

    if (row_num == 0):
        return list(worksheet.columns)
    else:
        row = []
        for c in range(worksheet.max_col):
            row.append(worksheet.iat[row_num - 1, c])
        return row


def read_workbook_col(worksheet_name, col_num, filename=FILE_NAME):
    """
    reads a single column of a worksheet from an Excel file and returns it as a list, with each list item
    representing one cell from the column
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param col_num: the number of the column to read (with the first column being column 1).
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """

    worksheet = open_if_not(filename)[worksheet_name]
    (max_row, max_col) = worksheet.shape

    col = [worksheet.columns[col_num]]
    for r in range(max_row):
        col.append(worksheet.iat[r, col_num])

    return clean_types(col, worksheet.columns[col_num])


def clean_types(column, col_name):
    """Takes a column of cells.  If one of the cells (excluding the
    header) is a time, assumes all values should be time, convents
    those that are to int and the rest to None with a warning.
    Otherwise, if any values are strings, it converts all non-None
    values to strings.  Otherwise, converts all non-None values to ints.
    """

    # Clean up column name.
    col_name = " ".join(col_name.split())

    # Add each type in the column to a set of types, excluding header.
    type_set = set()
    for i in range(1, len(column)):
        cell = column[i]
        type_set.add(type(cell))  
    #print("Before:", type_set)

    # If there's a time in the column, convert all values possible to ints, making the rest None.
    if datetime.time in type_set:
        for i in range(1, len(column)):
            if type(column[i]) == datetime.time:
                s = str(column[i])
                column[i] = int(s[0:2] + s[3:5] + s[6:8])
            elif column[i] is None:
                column[i] = None
            else:
                print((f"Warning: {column[i]} is invalid in time data, (column={col_name},"
                       f"row={i}), ignoring."), file=err_log)
                err_log.flush()
                column[i] = None
    # If there's a string in the column, convert all values to strings, unless None.
    elif str in type_set:  
        for i in range(1, len(column)):
            if column[i] is None:
                column[i] = None
            else:
                column[i] = str(column[i])
    # Everything else that isn't None should be ints.
    else:
        for i in range(1, len(column)):
            if column[i] is None:
                print(f"Warning: Missing entry in numeric data, (column={col_name}, row={i}).", file=err_log)
                err_log.flush()
                column[i] = None
            else:
                column[i] = int(column[i])

    # type_set = set()
    # for i in range(1, len(column)):
    #     cell = column[i]
    #     type_set.add(type(cell))  # add each type in the column to a set of types, excluding header
    # print("After:", type_set)
                
    return column


def read_workbook_cell(worksheet_name, row_num, col_num, filename=FILE_NAME):
    """
    reads a single cell of a worksheet from an Excel file and returns the contents of the cell.
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param row_num: the number of the row the desired cell is in (with the first row being row 1)
    @param col_num: the number of the column the desired cell is in (with the first column being column 1)
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """

    worksheet = open_if_not(filename)[worksheet_name]
    (max_row, max_col) = worksheet.shape
    return worksheet.iat[row_num - 1, col_num - 1]


def write_workbook(worksheet_name, data, filename=FILE_NAME):
    """
    Writes data to a worksheet in a given workbook
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param data: a list of lists representing the data to write to the spreadsheet.
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    NOTE: the indices for the list of lists start at zero, but they will be translated to Excel such that index zero
    is written to column/row 1
    """
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook[worksheet_name]
    
    for i in range(len(data)):
        x = i + 1
        column = data[i]
        for j in range(len(column)):
            y = j + 1
            worksheet.cell(row=x, column=y, value=column[j])
    workbook.save(filename)


def write_workbook_row(worksheet_name, row, data, filename=FILE_NAME):
    """
    Writes data to a given row in a given worksheet
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param row: the number of the row to write to (with the first row being row 1)
    @param data: a list containing the contents to write to each cell in the row
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook[worksheet_name]

    for i in range(len(data)):
        y = i + 1
        worksheet.cell(row=row + 1, column=y, value=data[i])
    workbook.save(filename)


def write_workbook_col(worksheet_name, col, data, filename=FILE_NAME):
    """
    Writes data to a given column in a given worksheet
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param col: the number of the column to write to (with the first column being column 1)
    @param data: a list containing the contents to write to each cell in the column
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook[worksheet_name]

    for i in range(len(data)):
        x = i + 1
        worksheet.cell(row=x, column=col + 1, value=data[i])
    workbook.save(filename)


def write_workbook_cell(worksheet_name, row, col, data, filename=FILE_NAME):
    """
    Writes data to a given cell in a given worksheet
    @param worksheet_name: a string representing the name of the worksheet within the workbook.
    @param row: the number of the row the target cell is in (starting with index 1)
    @param col: the number of the column the target cell is in (starting with index 1)
    @param data: the contents to write to this cell
    @param filename: the name of the file (i.e. workbook) to read from. Defaults to the global constant FILE_NAME.
    """
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook[worksheet_name]
    worksheet.cell(row=row + 1, column=col + 1, value=data)
    workbook.save(filename)

# def test_function_songj():
#     file = 'Copy of Class of 2026 (Fall 2022) Advisor and Registration Assignments - Anonymized.xlsx'
#     wb = openpyxl.load_workbook(filename=file)
#     ws = wb["Main Tab"]
#     dataset = []
#     for r in range(ws.max_row):
#         col = []
#         for l in range(ws.max_column):
#             col.append(ws.cell(r + 1, l + 1).value)
#         dataset.append(col)
#     pprint(dataset)
#
#     # data = [["a", "b", "c", "d", "e", "f"], ["g", "h", "i"], ["j", "k", "l", "m", "n"], ["o", "p", "q", "r", "s", "t"]]
#     data = [["A", "B", "C", "D", "E", "F"], ["G", "H", "I"], ["J", "K", "L", "M", "N"], ["O", "P", "Q", "R", "S", "T"]]
#     wb = openpyxl.Workbook()
#     ws = wb.create_sheet("test")
#
#     for r in range(len(data)):
#         x = r + 1
#         col = data[r]
#         for l in range(len(col)):
#             y = l + 1
#             ws.cell(row=x, column=y, value=col[l])
#     wb.save("test2.xlsx")
#     print("Excel created")


def letters_to_column_int(column_letters):
    if column_letters == "":
        return 0
    else:
        least_significant_char = column_letters[-1]
        remaining = column_letters[:-1]
        lsc_as_num = ord(least_significant_char) - 64
        full_num = letters_to_column_int(remaining) * 26 + lsc_as_num
        return full_num
        

# def test_read():
#     dataset = read_workbook("Main Tab")
#     for i in range(10):
#         print(dataset[i])
#
#     print()
#     print()
#     print(read_workbook_row("Main Tab", 1))
#
#     print()
#     print()
#     print(read_workbook_col("Main Tab", 1))
#
#
# def test_write():
#     filename = "Test_Write.xlsx"
#
#     sheet1 = [["A", "B", "C"], ["a", "b", "c"], ["1", "2", "3"]]
#     sheet2 = ["a", "b", "c"]
#     sheet3 = ["B", "b", "2"]
#     sheet4 = "b"
#
#     write_workbook("Sheet1", sheet1, filename)
#     write_workbook_row("Sheet2", 1, sheet2, filename)
#     write_workbook_col("Sheet3", 1, sheet3, filename)
#     write_workbook_cell("Sheet4", 1, 1, sheet4, filename)


if __name__ == '__main__':
    # column_names = read_workbook_col("Main Tab", 0)
    workbook = read_workbook("Main Tab")
    # print(workbook)
    
    
    
